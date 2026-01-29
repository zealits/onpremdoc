"""Dump Document_Processing_Pipeline_Steps.xlsx to text for inspection."""
import openpyxl
from pathlib import Path

path = Path("Document_Processing_Pipeline_Steps.xlsx")
if not path.exists():
    print("File not found:", path.absolute())
    exit(1)

wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
out_lines = []
for i, sheet_name in enumerate(wb.sheetnames):
    out_lines.append(f"=== Sheet {i+1}: {sheet_name} ===")
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        out_lines.append(repr(tuple(c for c in row)))
        if row_idx >= 50:
            out_lines.append("... (truncated)")
            break
    out_lines.append("")
wb.close()

out_path = Path("pipeline_excel_dump.txt")
out_path.write_text("\n".join(out_lines), encoding="utf-8")
print("Wrote", out_path)
