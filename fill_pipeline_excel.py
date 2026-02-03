"""
Fill Document_Processing_Pipeline_Steps.xlsx:
- Pipeline Steps: Step #, Step Name, Description, Module/File, Models & Technologies,
  Algorithms & Methods, Key Functions, Input, Output, Examples.
- Summary: Remove re-ranking; no mention of reranking (pipeline does not use it).
- Combined Steps sheet: same columns, fewer rows; formatted like Pipeline Steps (wrap, colors, spacing).
"""
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill

# Step#, Step Name, Module/File, Models & Technologies, Algorithms & Methods, Key Functions (one purpose per step)
STEPS = [
    (1, "Upload PDF", "main.py", "FastAPI", "File save, UUID generation", "upload_pdf(), File(), UploadFile"),
    (2, "Schedule PDF processing task", "main.py, detection.py", "—", "Enqueue background task that runs steps 3–6", "process_pdf_background(), process_single_pdf()"),
    (3, "Convert PDF to markdown", "detection.py", "Docling (DocumentConverter)", "PDF parse, export with page-break placeholder", "DocumentConverter.convert(), export_to_markdown(page_break_placeholder)"),
    (4, "Correct heading hierarchy", "detection.py", "docling-hierarchical (ResultPostprocessor)", "TOC/bookmark level inference", "ResultPostprocessor(doc, source=pdf_path).process()"),
    (5, "Fix markdown tables", "detection.py", "—", "Table detect, separator/header/column fix", "process_markdown(), fix_table(), trim_trailing_empty_columns()"),
    (6, "Extract page mapping", "detection.py", "—", "Scan page-break markers, assign line→page", "extract_page_mapping_from_markdown(), create_approximate_page_mapping()"),
    (7, "Schedule vectorization task", "main.py", "FastAPI, LangGraph", "Enqueue workflow that runs steps 8–19", "vectorize_document(), vectorize_background(), create_vectorization_workflow()"),
    (8, "Load markdown and parse", "vectorizerE.py", "—", "Read file, parse lines/blocks, attach page numbers", "load_markdown(), parse_markdown_enhanced()"),
    (9, "Extract document structure", "vectorizerE.py", "—", "Regex headers, build section tree with full paths", "extract_document_structure(), header match (#+), section path tree"),
    (10, "Create enhanced chunks", "vectorizerE.py", "—", "Split by size/overlap, assign section/page/prev/next", "create_enhanced_chunk(), is_chunk_empty(), adjacency metadata"),
    (11, "Initialize LLM and vector store", "vectorizerE.py", "Ollama (llama3.1:8b), nomic-embed-text:v1.5, Chroma", "Embedding model, LLM, persistent vector DB", "Ollama(), OllamaEmbeddings(), Chroma(persist_directory)"),
    (12, "Build section graph", "vectorizerE.py", "NetworkX", "Section nodes, parent-child edges (contains_section)", "add_section_node(), add_edge(relation='contains_section')"),
    (13, "Page classification", "vectorizerE.py", "Ollama LLM", "One label per page (1–4 words)", "classify_pages(), classify_page_with_llm()"),
    (14, "Process each chunk (steps 15–18)", "vectorizerE.py", "Ollama, Chroma, NetworkX", "Loop chunks; run 15 (summary), 16 (embed), 17 (edges), 18 (similar_to)", "process_chunks_one_by_one()"),
    (15, "Generate chunk summary", "vectorizerE.py", "Ollama LLM", "One-line summary, token limit", "get_summary_from_llm(), token_tracker.check_llm_limit()"),
    (16, "Embed chunk and add to Chroma", "vectorizerE.py", "Ollama (nomic-embed-text), Chroma", "Truncate if over limit; embed; add_documents()", "token_tracker.check_embedding_limit(), vector_store.add_documents()"),
    (17, "Add chunk node and section/page edges", "vectorizerE.py", "NetworkX", "One node per chunk; edges: belongs_to, on_page, follows", "add_chunk_node(), add_edge(belongs_to, on_page, follows)"),
    (18, "Add similar_to edges", "vectorizerE.py", "Chroma, NetworkX", "Vector search from chunk; threshold; add similar_to edges", "similarity_search_with_score(), add_edge(relation='similar_to')"),
    (19, "Persist graph, mapping, Chroma", "vectorizerE.py", "Chroma, JSON", "Write graph JSON, mapping JSON, Chroma to disk", "document_graph.save(), json.dump(), persist_directory"),
    (20, "Submit query and invoke agent", "main.py", "FastAPI", "Validate request, load agent, invoke graph", "query_document(), load_agent_for_document(), agent.invoke()"),
    (21, "Load agent resources", "main.py, retrivalAgentE.py", "Chroma, NetworkX, Ollama", "Load vector store, graph, chunks, LLM, page agent", "load_chunks_from_mapping(), load_vector_store(), DocumentGraph.load(), load_page_agent()"),
    (22, "Classify query type", "retrivalAgentE.py", "Ollama LLM, regex", "Page-summary vs normal retrieval; route", "classify_query(), route_query_type()"),
    (23, "Vector similarity search", "retrivalAgentE.py", "Chroma, nomic-embed-text", "Embed query; top-k by L2; distance→similarity", "similarity_search_with_score(query, k=20)"),
    (24, "Graph expansion", "retrivalAgentE.py", "NetworkX", "From top seeds: sections, adjacent, similar_to; cap", "expand_from_chunks(), get_parent_section(), get_adjacent_chunks(), get_similar_chunks()"),
    (25, "Cap and set retrieved chunks", "retrivalAgentE.py", "—", "Take first N chunks (e.g. 25); store in state", "retrieved_chunks[:25], state['retrieved_chunks']"),
    (26, "Analyze chunks; decide if more info needed", "retrivalAgentE.py", "Ollama LLM", "LLM: suffice? one follow-up query if not", "analyze_chunks(), should_continue_search()"),
    (27, "Second retrieval (if needed)", "retrivalAgentE.py", "Chroma, NetworkX", "Vector search + expand with new_query; cap; append", "second_retrieval(), similarity_search_with_score(new_query, k=15)"),
    (28, "Generate final answer", "retrivalAgentE.py", "Ollama LLM", "Prompt: question + chunks; LLM answer from chunks only", "generate_final_answer(), answer_prompt with QUESTION"),
    (29, "Format API response", "main.py", "FastAPI, Pydantic", "Map state to QueryResponse schema", "QueryResponse(), ChunkDetail(), retrieval_stats"),
    (30, "Page summarization", "main.py, page_summarization.py", "Ollama LLM, DocumentGraph", "Load page chunks; LLM summary and key points", "load_page_agent(), summarize_page(), _generate_summary_with_llm()"),
]

# One purpose per step; no inheritance or forward/backward references. Column C.
DESCRIPTIONS = {
    1: "Receive the PDF file from the client and save it to a document folder named by a new UUID. Return the document_id and status so the client can poll or call vectorize/query.",
    2: "Enqueue a background task for this document. When the task runs, it executes steps 3, 4, 5, and 6 in order and writes the markdown and page-mapping files to the document folder. The API returns immediately.",
    3: "Convert the PDF to structured markdown using Docling. Insert page-break placeholders in the output. Output is raw markdown (headings, lists, tables, page-break markers).",
    4: "Adjust heading levels in the document using the PDF TOC or bookmarks so that section nesting is correct (e.g. Part A as top level, Section 1.2 as child). Output is the same document with corrected heading levels.",
    5: "Detect tables in the markdown and fix separator rows, header rows, and trailing empty columns. Output is markdown with valid table syntax.",
    6: "Scan the markdown for page-break markers and assign a page number to each line or block. Output is a JSON map (line_to_page, page_boundaries, total_pages).",
    7: "Enqueue a background vectorization workflow for the document. Create the workflow state (markdown path, document folder) and run the workflow. The API returns immediately.",
    8: "Read the markdown file and page-mapping JSON from disk. Parse the markdown into lines or blocks and attach a page number to each using the page mapping. Output is parsed content with page numbers.",
    9: "Scan the markdown for headers (#, ##, etc.) and build a tree of sections. Assign each section a full path (e.g. 'Part A > Section 2 > 2.1'). Output is a structure dict (sections with full paths).",
    10: "Split the markdown into fixed-size chunks (e.g. 2000 chars, 150 overlap). For each chunk, set metadata: section_path from structure, page_number from page mapping, heading, prev/next chunk IDs. Drop empty or duplicate chunks. Output is a list of Document objects with metadata.",
    11: "Create the Ollama LLM and embedding client and a Chroma vector store with a persist directory. Output is initialized LLM, embeddings, and vector_store used in steps 15–16 and by the retrieval agent.",
    12: "Create one graph node per section from the structure tree. Add a directed edge from each parent section to its child. Output is a graph with section nodes and contains_section edges.",
    13: "For each page, call the LLM to assign a short label (1–4 words) from the page content. Output is a dict mapping page_number to classification label.",
    14: "For each chunk in order: run step 15 (summary), step 16 (embed and add to Chroma), step 17 (add chunk node and section/page/prev/next edges), step 18 (add similar_to edges). Output is updated vector store, graph, and JSON mapping entries.",
    15: "Call the LLM with the chunk text to produce one concise summary line (token-limited). Attach the summary to the chunk metadata. Output is the chunk with metadata.summary set.",
    16: "Truncate the chunk text if it exceeds embedding limits. Compute the embedding and add the chunk (with metadata) to the Chroma collection. Output is the chunk stored in Chroma.",
    17: "Add one graph node for the chunk. Add edges: chunk→section (belongs_to), chunk→page (on_page), chunk→prev (follows), chunk→next (follows). Output is the updated graph.",
    18: "Run a vector search from this chunk’s embedding; find other chunks within the similarity threshold. Add a similar_to edge from this chunk to each. Output is the updated graph.",
    19: "Write the graph to a JSON file. Write the vector mapping (chunk list with metadata) to a JSON file. Persist the Chroma collection to disk. Output is the three artifacts on disk.",
    20: "Validate the request (query, document_id). Load the agent resources for that document. Put the query into the agent state and invoke the agent graph. Output is the initial state and graph invocation.",
    21: "Load the Chroma vector store from the document’s vector_db path. Load the document graph from the graph JSON. Load the chunk list from the vector mapping JSON. Load the LLM and, if configured, the page summarization agent. Output is these objects in memory for the retrieval nodes.",
    22: "Decide whether the query asks for a page summary (e.g. 'summarize page 5') or a normal retrieval. Use pattern matching and optionally the LLM. Output is is_page_summary and, if true, page_number. This routes to either step 30 or step 23.",
    23: "Embed the query and run a vector search in Chroma (top-k by L2 distance). Convert distances to similarity scores. Output is the list of seed chunk IDs and their similarity scores.",
    24: "Starting from the top seed chunks (e.g. top 5), traverse the graph: add parent sections, adjacent chunks (prev/next), and similar_to neighbours. Stop after a fixed number of extra chunks (e.g. 15). Output is the expanded set of chunk IDs.",
    25: "Take the union of seed and expanded chunk IDs and keep only the first N (e.g. 25) chunks as Document objects. Store them in state. Output is state['retrieved_chunks'].",
    26: "Call the LLM with the query and a subset of the retrieved chunks. Ask whether the chunks suffice to answer the query and, if not, for one follow-up query. Output is needs_more_info and, if true, new_query.",
    27: "If needs_more_info and new_query exist: embed new_query, run vector search (top-k), expand via graph, exclude already-retrieved chunks, cap the new set (e.g. 15). Append to state. Output is state['second_retrieval_chunks'].",
    28: "Build a prompt that states the user question at the top and again before the answer instruction. Fill the middle with the selected chunks (primary and, if any, second). Call the LLM. Return only the answer text; do not cite chunk numbers. Output is state['final_answer'].",
    29: "Map the agent state (final_answer, chunks, retrieval stats) into the API response schema: answer, optional chunk details, retrieval_stats, second_query if used. Return the JSON response.",
    30: "Load chunks for the requested page (and optionally adjacent pages) from the graph/mapping. Call the LLM to produce a short summary, key points, and section labels. Return this as the final answer.",
}

# Input (column H) and Output (column I) — only what this step consumes/produces
INPUT_OUTPUT = {
    1: ("PDF file (UploadFile)", "Saved PDF path, document_id, ProcessPDFResponse"),
    2: ("document_id, PDF path", "Enqueued task (steps 3–6 run when task executes)"),
    3: ("PDF file", "Raw markdown with page break markers"),
    4: ("Docling document object", "Document with corrected heading hierarchy"),
    5: ("Markdown text", "Fixed markdown (valid tables)"),
    6: ("Markdown with page breaks", "Page mapping JSON: {line_to_page, page_boundaries, total_pages}"),
    7: ("document_id", "Enqueued workflow (steps 8–19 run when workflow runs)"),
    8: ("Markdown file path, page mapping path", "Parsed content (lines/blocks) with page numbers"),
    9: ("Markdown content", "Structure dict: sections with full paths"),
    10: ("Markdown, structure, page mapping", "List of Document chunks with metadata"),
    11: ("Model names, base URL", "LLM, embeddings, vector_store, DocumentGraph (empty)"),
    12: ("Structure dict", "Graph with section nodes and contains_section edges"),
    13: ("Chunks by page, page mapping", "{page_number: classification_label}"),
    14: ("Chunk list, LLM, vector_store, graph", "Updated Chroma, graph, JSON mapping (via 15–18)"),
    15: ("Chunk text", "Chunk with metadata.summary"),
    16: ("Chunk (text + metadata)", "Chunk stored in Chroma"),
    17: ("Chunk, graph, structure", "Graph with chunk node and belongs_to/on_page/follows edges"),
    18: ("Chunk embedding, Chroma, graph", "Graph with similar_to edges from this chunk"),
    19: ("DocumentGraph, JSON mapping, Chroma", "Graph JSON file, vector_mapping JSON file, Chroma on disk"),
    20: ("Query string, document_id", "Agent state after graph invocation"),
    21: ("document_id", "Loaded vector_store, graph, chunks, LLM, page agent"),
    22: ("User query", "is_page_summary, page_number (if page summary); route"),
    23: ("Query string", "Seed chunk IDs, similarity scores"),
    24: ("Seed chunk IDs", "Expanded chunk IDs (seed + graph neighbours)"),
    25: ("Seed + expanded chunk IDs, chunk list", "state['retrieved_chunks'] (first N)"),
    26: ("Query, retrieved chunks", "needs_more_info, new_query (optional)"),
    27: ("new_query, existing chunk IDs", "state['second_retrieval_chunks'] (capped)"),
    28: ("Query, primary chunks, second chunks", "state['final_answer']"),
    29: ("Agent state (final_answer, chunks, stats)", "QueryResponse JSON"),
    30: ("Page number", "Page summary text, key points, sections"),
}

# Simple examples. Column J. One action per step.
EXAMPLES = {
    1: "POST /upload with policy.pdf → file saved to output/{uuid}/policy.pdf → response: { document_id, status: 'processing' }.",
    2: "process_pdf_background(pdf_path, document_id) called → task enqueued; when it runs, it will execute steps 3, 4, 5, 6.",
    3: "Docling converts policy.pdf → raw markdown with '## Part A', '<!-- page break -->', tables.",
    4: "ResultPostprocessor(doc).process() → headings adjusted using PDF bookmarks (e.g. # Part A, ## Section 1).",
    5: "process_markdown(markdown) → tables fixed (separators, headers, empty columns).",
    6: "extract_page_mapping_from_markdown(md) → { line_to_page: {0:1, 45:2, ...}, total_pages: 10 }.",
    7: "POST /vectorize { document_id } → vectorize_background(document_id) enqueued; workflow will run steps 8–19.",
    8: "load_markdown(path), page_mapping → parsed lines/blocks, each with page_number.",
    9: "extract_document_structure(md) → { sections: [ { path: 'Part A > 2.1', ... } ] }.",
    10: "create_enhanced_chunk() in loop → list of Document chunks with section_path, page_number, heading, prev/next.",
    11: "Ollama(), OllamaEmbeddings(), Chroma(persist_directory=...) → LLM, embeddings, vector_store ready.",
    12: "For each section in structure: add_section_node(), add_edge(parent, child, 'contains_section').",
    13: "classify_pages() → { 1: 'Cover', 2: 'Terms', 3: 'Terms', ... }.",
    14: "For each chunk: call step 15, then 16, then 17, then 18 → Chroma and graph updated.",
    15: "get_summary_from_llm(chunk_text) → 'Clause stating governing law is India.'; store in chunk.metadata.",
    16: "vector_store.add_documents([doc]) → chunk embedded and stored in Chroma.",
    17: "add_chunk_node(chunk); add_edge(chunk, section, 'belongs_to'); add_edge(chunk, page, 'on_page'); add_edge(chunk, prev, 'follows').",
    18: "similarity_search_with_score(chunk_embedding) → add similar_to edges to chunks within threshold.",
    19: "document_graph.save('graph.json'); json.dump(mapping, 'vector_mapping.json'); Chroma persists.",
    20: "POST /query { document_id, query } → load_agent_for_document(); agent.invoke(initial_state) → final_state.",
    21: "load_chunks_from_mapping(), load_vector_store(), DocumentGraph.load() → vector_store, graph, chunks, LLM in memory.",
    22: "classify_query(state) → is_page_summary=True, page_number=5 → route to step 30; else → step 23.",
    23: "similarity_search_with_score(query, k=20) → seed_chunk_ids=[7,12,...], seed_chunk_scores={7:0.82,...}.",
    24: "expand_from_chunks(top_5_seeds) → graph_expanded_ids (e.g. 16 more chunk IDs).",
    25: "retrieved_chunks = [doc for id in (seed_ids + expanded_ids)][:25] → state['retrieved_chunks'].",
    26: "analyze_chunks(state) → needs_more_info=True, new_query='Which country laws govern this policy?'.",
    27: "second_retrieval(state) → vector search + expand with new_query → state['second_retrieval_chunks'] = [8 chunks].",
    28: "generate_final_answer(state) → prompt with QUESTION + chunks → LLM → state['final_answer'] = answer text.",
    29: "QueryResponse(answer=final_state['final_answer'], chunks_detail=..., retrieval_stats=...) → JSON response.",
    30: "summarize_page(state) → load page 5 chunks → LLM → 'Page 5: Premium payment. Key points: ...'.",
}

# --------------- COMBINED STEPS (new sheet): fewer steps, all details preserved ---------------
# Each tuple: (step#, name, description, module_file, models_tech, algorithms, key_functions, input, output, examples)
COMBINED_STEPS = [
    (1, "Upload & schedule PDF processing", "Receive PDF from client; save to document folder (UUID); return document_id and status. Enqueue a background task that, when run, executes: convert PDF to markdown (Docling, page-break placeholders), correct heading hierarchy (ResultPostprocessor from TOC/bookmarks), fix markdown tables (process_markdown), extract page mapping from markers. Task writes .md file and _page_mapping.json to the document folder.", "main.py, detection.py", "FastAPI, Docling, docling-hierarchical", "File save, UUID, background enqueue; PDF→markdown, heading correction, table fix, line→page extraction", "upload_pdf(), process_pdf_background(), process_single_pdf(), DocumentConverter.convert(), ResultPostprocessor.process(), process_markdown(), extract_page_mapping_from_markdown()", "PDF file (UploadFile)", "Saved PDF path, document_id, ProcessPDFResponse; then (when task runs) markdown file, page mapping JSON", "POST /upload → file saved, task enqueued; task runs: Docling→md, headings fixed, tables fixed, page_mapping.json written."),
    (2, "PDF to markdown pipeline (convert, headings, tables, page mapping)", "Convert PDF to structured markdown with page-break placeholders (Docling). Correct heading levels using PDF TOC/bookmarks (ResultPostprocessor). Fix tables: separators, headers, trailing columns (process_markdown). Scan markers and build line-to-page map (extract_page_mapping_from_markdown or create_approximate_page_mapping). Save fixed markdown and page mapping JSON. All details as in steps 3–6.", "detection.py", "Docling, docling-hierarchical", "PDF parse, export_to_markdown(page_break_placeholder), TOC heading inference, table detect/fix, page-break scan", "DocumentConverter.convert(), export_to_markdown(), ResultPostprocessor.process(), process_markdown(), fix_table(), extract_page_mapping_from_markdown(), create_approximate_page_mapping()", "PDF file", "Raw markdown with page breaks; document with corrected headings; fixed markdown; page mapping JSON {line_to_page, page_boundaries, total_pages}. Files written to document folder.", "policy.pdf → Docling → md with breaks; headings fixed; tables fixed; extract_page_mapping_from_markdown() → JSON; both files saved."),
    (3, "Schedule vectorization & load document (parse, structure)", "Enqueue background vectorization workflow; create state (markdown path, document folder). When workflow runs: load markdown file and page-mapping JSON; parse markdown into lines/blocks and attach page numbers. Scan markdown for headers (#, ##, …) and build section tree with full paths (e.g. 'Part A > Section 2 > 2.1'). Combines steps 7–9.", "main.py, vectorizerE.py", "FastAPI, LangGraph", "Background enqueue; read file, parse, attach page numbers; regex headers, section tree with full paths", "vectorize_document(), vectorize_background(), load_markdown(), parse_markdown_enhanced(), extract_document_structure()", "document_id; then markdown path, page mapping path", "Enqueued workflow; then parsed content (lines/blocks + page numbers), structure dict (sections with full paths)", "POST /vectorize → workflow enqueued; workflow: load_markdown(), parse → blocks+page; extract_document_structure() → section tree."),
    (4, "Create chunks & initialize models", "Split markdown into fixed-size chunks (e.g. 2000 chars, 150 overlap). Assign each chunk: section_path from structure, page_number from page mapping, heading, prev/next chunk IDs; drop empty/duplicate. Initialize Ollama LLM and embedding model; create Chroma vector store with persist directory. Combines steps 10–11.", "vectorizerE.py", "Ollama (llama3.1, nomic-embed-text), Chroma", "Size/overlap split, section/page/prev/next metadata; LLM and embedding client, Chroma persist_directory", "create_enhanced_chunk(), is_chunk_empty(); Ollama(), OllamaEmbeddings(), Chroma(persist_directory)", "Markdown, structure, page mapping; model names, base URL", "List of Document chunks with metadata; initialized LLM, embeddings, vector_store, empty DocumentGraph", "Chunks created with section_path, page_number; Ollama(), Chroma() created."),
    (5, "Build section graph & classify pages", "Add one graph node per section from structure; add directed edge from each parent to child (contains_section). For each page, call LLM to assign a short label (1–4 words) from page content. Output: graph with section nodes and edges; dict page_number → classification_label. Combines steps 12–13.", "vectorizerE.py", "NetworkX, Ollama LLM", "Section nodes, parent→child edges; per-page LLM classification", "add_section_node(), add_edge(contains_section); classify_pages(), classify_page_with_llm()", "Structure dict; chunks grouped by page, page mapping", "Graph with section nodes and contains_section edges; {page_number: classification_label}", "add_section_node() for each section; classify_pages() → {1:'Cover', 2:'Terms', ...}."),
    (6, "Process chunks: summary, embed, graph (per chunk)", "For each chunk in order: (1) Generate one-line summary via LLM (token-limited); attach to chunk metadata. (2) Truncate if over embedding limit; compute embedding; add chunk to Chroma. (3) Add chunk node to graph; edges: belongs_to section, on_page, follows prev/next. (4) Vector search from chunk embedding; add similar_to edges to chunks within threshold. Combines steps 14–18.", "vectorizerE.py", "Ollama LLM, nomic-embed-text, Chroma, NetworkX", "Per chunk: get_summary_from_llm; add_documents; add_chunk_node, belongs_to/on_page/follows; similarity_search_with_score, similar_to edges", "process_chunks_one_by_one(), get_summary_from_llm(), vector_store.add_documents(), add_chunk_node(), add_edge(), similarity_search_with_score()", "Chunk list, LLM, vector_store, graph, structure", "Chroma filled with chunks; graph with chunk nodes, section/page/follows/similar_to edges; JSON mapping entries", "For each chunk: summary → embed → Chroma; chunk node + edges; similar_to edges from vector search."),
    (7, "Persist vectorization outputs", "Write document graph to JSON file. Write vector mapping (chunk list with metadata) to JSON file. Persist Chroma collection to disk. Retrieval agent will load these. Step 19.", "vectorizerE.py", "Chroma, JSON", "Save graph JSON, dump mapping JSON, Chroma persist_directory", "document_graph.save(), json.dump(), Chroma persist", "DocumentGraph, JSON mapping, Chroma", "Graph JSON file, vector_mapping JSON file, Chroma on disk", "document_graph.save('graph.json'); json.dump(mapping, 'vector_mapping.json'); Chroma persists."),
    (8, "Submit query & load agent", "Validate request (query, document_id). Load agent resources: Chroma vector store, document graph (NetworkX), chunk list from vector mapping JSON, LLM; optionally page summarization agent. Put query into agent state and invoke the agent graph. Combines steps 20–21.", "main.py, retrivalAgentE.py", "FastAPI, Chroma, NetworkX, Ollama", "Request validation; load vector store, graph, chunks, LLM, page agent; agent.invoke(initial_state)", "query_document(), load_agent_for_document(), load_chunks_from_mapping(), load_vector_store(), DocumentGraph.load(), load_page_agent(), agent.invoke()", "Query string, document_id", "Agent state after graph invocation (includes loaded vector_store, graph, chunks, LLM)", "POST /query → load_agent_for_document(); agent.invoke({query, ...}) → final_state."),
    (9, "Classify query & retrieve (vector search + graph expansion)", "Classify query as page-summary (e.g. 'summarize page 5') or normal retrieval; route accordingly. For normal: embed query; vector search in Chroma (top-k by L2 distance); convert to similarity scores. From top seed chunks, traverse graph (parent sections, adjacent prev/next, similar_to); cap expansion. Take first N chunks (e.g. 25) as retrieved_chunks. Combines steps 22–25.", "retrivalAgentE.py", "Ollama LLM, regex, Chroma, nomic-embed-text, NetworkX", "classify_query, route; similarity_search_with_score(query, k=20); expand_from_chunks (sections, adjacent, similar_to); cap N chunks", "classify_query(), route_query_type(), similarity_search_with_score(), distance_to_similarity(), expand_from_chunks(), get_parent_section(), get_adjacent_chunks(), get_similar_chunks()", "User query; then query string; then seed chunk IDs", "is_page_summary, page_number (if page); or seed_chunk_ids, similarity scores; expanded IDs; state['retrieved_chunks']", "classify → normal_retrieval; vector search → seeds; expand_from_chunks() → state['retrieved_chunks'][:25]."),
    (10, "Analyze chunks & optional second retrieval", "LLM analyzes retrieved chunks: do they suffice to answer the query? If not, produce one follow-up query. If needs_more_info and new_query: embed new_query; vector search (top-k); graph expand; exclude already-retrieved; cap new set (e.g. 15); append to state as second_retrieval_chunks. Combines steps 26–27.", "retrivalAgentE.py", "Ollama LLM, Chroma, NetworkX", "analyze_chunks (suffice? one follow-up query); second_retrieval: vector search + expand, cap, append", "analyze_chunks(), should_continue_search(), second_retrieval(), similarity_search_with_score(new_query, k=15)", "Query, retrieved chunks; then new_query, existing chunk IDs", "needs_more_info, new_query; state['second_retrieval_chunks'] (if second retrieval run)", "analyze_chunks() → needs_more_info=True, new_query='...'; second_retrieval() → 8 new chunks appended."),
    (11, "Generate final answer & format API response", "Build prompt with user question at top and repeated before answer instruction; fill context with primary and (if any) second retrieval chunks. Call LLM; answer from chunks only; no chunk numbers in answer. Map agent state to API schema: answer, optional chunk details, retrieval_stats, second_query if used. Combines steps 28–29.", "retrivalAgentE.py, main.py", "Ollama LLM, FastAPI, Pydantic", "generate_final_answer (QUESTION + chunks prompt); QueryResponse schema mapping", "generate_final_answer(), answer_prompt with QUESTION block; QueryResponse(), ChunkDetail(), retrieval_stats", "Query, primary chunks, second chunks; then final_answer, chunks, stats", "state['final_answer']; QueryResponse JSON (answer, chunks_detail, retrieval_stats)", "generate_final_answer() → prompt + LLM → final_answer; QueryResponse(...) → JSON response."),
    (12, "Page summarization", "When query is classified as page-summary: load chunks for the requested page (and optionally adjacent pages) from graph/mapping. Call LLM to produce a short summary, key points, and section labels. Return as final answer. Step 30.", "main.py, page_summarization.py", "Ollama LLM, DocumentGraph", "Load page chunks (and adjacent if needed); LLM summary and key points", "load_page_agent(), summarize_page(), _generate_summary_with_llm()", "Page number", "Page summary text, key points, sections, classification", "Query 'Summarize page 5' → load page 5 chunks → LLM → 'Page 5: Premium payment. Key points: ...'."),
]


def main():
    path = Path("Document_Processing_Pipeline_Steps.xlsx")
    if not path.exists():
        print("File not found:", path.absolute())
        return

    wb = openpyxl.load_workbook(path)
    if "Pipeline Steps" not in wb.sheetnames:
        print("Sheet 'Pipeline Steps' not found")
        return

    ws = wb["Pipeline Steps"]
    # Add Examples header if column 10 not present
    if ws.max_column < 10:
        ws.cell(row=1, column=10, value="Examples")
    else:
        ws.cell(row=1, column=10, value="Examples")

    for i, row_data in enumerate(STEPS):
        step_num, step_name, module_file, models_tech, algorithms, key_functions = row_data
        row_idx = i + 2
        ws.cell(row=row_idx, column=1, value=step_num)
        ws.cell(row=row_idx, column=2, value=step_name)
        ws.cell(row=row_idx, column=3, value=DESCRIPTIONS.get(step_num, ""))
        ws.cell(row=row_idx, column=4, value=module_file)
        ws.cell(row=row_idx, column=5, value=models_tech)
        ws.cell(row=row_idx, column=6, value=algorithms)
        ws.cell(row=row_idx, column=7, value=key_functions)
        inp_out = INPUT_OUTPUT.get(step_num, ("", ""))
        ws.cell(row=row_idx, column=8, value=inp_out[0])
        ws.cell(row=row_idx, column=9, value=inp_out[1])
        ws.cell(row=row_idx, column=10, value=EXAMPLES.get(step_num, ""))

    # Summary sheet: remove any mention of re-ranking (pipeline does not use it)
    if "Summary" in wb.sheetnames:
        sum_ws = wb["Summary"]
        for row_idx in range(1, sum_ws.max_row + 1):
            val = sum_ws.cell(row=row_idx, column=1).value
            if val and "re-ranking" in str(val).lower():
                sum_ws.cell(row=row_idx, column=1).value = "- Chunk selection (top-k from vector + graph)"
                if sum_ws.cell(row=row_idx, column=2).value:
                    sum_ws.cell(row=row_idx, column=2).value = None
                break

    # New sheet: Combined Steps (fewer steps, all details preserved) — same visual style as Pipeline Steps
    combined_sheet_name = "Combined Steps"
    if combined_sheet_name in wb.sheetnames:
        del wb[combined_sheet_name]
    combined_ws = wb.create_sheet(combined_sheet_name)
    headers = ("Step #", "Step Name", "Description", "Module/File", "Models & Technologies", "Algorithms & Methods", "Key Functions", "Input", "Output", "Examples")
    for col, h in enumerate(headers, 1):
        combined_ws.cell(row=1, column=col, value=h)
    for i, row_data in enumerate(COMBINED_STEPS):
        row_idx = i + 2
        for col, val in enumerate(row_data, 1):
            combined_ws.cell(row=row_idx, column=col, value=val)

    # Match Pipeline Steps formatting: header row (blue fill, white bold, center, wrap); data (wrap, top align); column widths; row heights
    header_fill = PatternFill(patternType="solid", fgColor="FF366092", bgColor="FF366092")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, 11):
        combined_ws.cell(row=1, column=col).fill = header_fill
        combined_ws.cell(row=1, column=col).font = header_font
        combined_ws.cell(row=1, column=col).alignment = header_alignment
    for row_idx in range(2, combined_ws.max_row + 1):
        for col in range(1, 11):
            combined_ws.cell(row=row_idx, column=col).alignment = data_alignment
    # Same column widths as Pipeline Steps
    combined_ws.column_dimensions["A"].width = 8.0
    combined_ws.column_dimensions["B"].width = 25.0
    combined_ws.column_dimensions["C"].width = 50.0
    combined_ws.column_dimensions["D"].width = 20.0
    combined_ws.column_dimensions["E"].width = 35.0
    combined_ws.column_dimensions["F"].width = 40.0
    combined_ws.column_dimensions["G"].width = 30.0
    combined_ws.column_dimensions["H"].width = 25.0
    combined_ws.column_dimensions["I"].width = 30.0
    combined_ws.column_dimensions["J"].width = 38.29
    # Row heights: header; data rows taller so wrapped text fits
    combined_ws.row_dimensions[1].height = 15.75
    for row_idx in range(2, combined_ws.max_row + 1):
        combined_ws.row_dimensions[row_idx].height = 90

    wb.save(path)
    print("Filled Pipeline Steps + Combined Steps sheet. Saved:", path.absolute())


if __name__ == "__main__":
    main()
